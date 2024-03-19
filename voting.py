"""
In this project, I have designed and implement several voting rules.
"""
import openpyxl

def generate_preferences(values):
    """generate_preferences function takes in a worksheet of numerical values representing the various alternatives for agents and
       produces a preference profile as its output."""

    values_list = list()
    for row in values.iter_rows(min_row=1, min_col=1, values_only=True):
        values_list.append(row)  # Appending each row values in value_list list
    preference_profile = {}

    for key, lists in enumerate(values_list, start=1):
        # Iterating through each values and creating a preference list based on valuations

        index = list(range(len(lists)))
        preference_list = []
        while index:
            max_index = 0
            for i in range(1, len(index)):
                if lists[index[i]] >= lists[index[max_index]]:
                    max_index = i
            preference_list.append(index.pop(max_index)+1)

        preference_profile[key] = preference_list

    return preference_profile

def dictatorship(preferences, agent):
    """dictatorship function takes in a preference profile and an integer corresponding to an agent, and
       produces the winner according to the Dictatorship rule as its output."""

    if agent not in preferences:
        raise ValueError("Entered agent is invalid!!")  # Check if the agent exists in the preference profile and raise error message if not

    temp = preferences[agent]
    agent_preference = temp[0]
    # Get the preferences of the agent and return the alternative that the agent ranks first
    return agent_preference

def scoring_rule(preferences, score_vector, tie_break):
    """scoring_rule function takes in a preference profile, a score vector and an integer corresponding to an agent,
       and produces the winner according to the scroring rule as its output."""

    length = len(score_vector)

    # Checking the length of score_vector is same as the number of alternatives, if not prints message and return false
    if length != len(preferences[1]):
        print("Incorrect input")
        return False

    score = {}
    for i in range(1, length + 1):
        score[i] = 0

    score_vector.sort(reverse=True)
    # Assigning score as per preferences and a vector of scores.
    for value in preferences.values():
        for i, j in enumerate(value):
            score[j] += score_vector[i]

    max_score = max(score.values())
    winners = find_winner(score, max_score)  # Calling find_winner function to find winners with max score

    return tie_breaking(winners, preferences, tie_break)  # Calling tie_breaking function to find a single winner if there are multiple

def plurality(preferences, tie_break):
    """plurality function takes in a preference profile and an integer corresponding to an agent,
       and produces the winner according to the plurality rule as its output."""

    count = {}
    for value in preferences.values():
        first = value[0]
        if first in count:
            count[first] += 1
        else:
            count[first] = 1

    max_count = max(count.values())
    winners = find_winner(count, max_count)  # Find the winners with the maximum count by calling find_winner

    return tie_breaking(winners, preferences, tie_break)  # Find a single winner if multiple are there, by calling tie_breaking

def veto(preferences, tie_break):
    """veto function takes in a preference profile and an integer corresponding to an agent,
       and produces the winner according to the veto rule as its output."""

    length = len(preferences[1])

    score = {}
    for i in range(1, length + 1):
        score[i] = 0  # Assigning score 0 to each alternatives

    for value in preferences.values():
        for i, j in enumerate(value):
            if i != length - 1:  # Assigning score 1 to alternatives not in last position
                score[j] += 1

    max_score = max(score.values())
    winners = find_winner(score, max_score)  # Find the winners with the maximum score by calling find_winner

    return tie_breaking(winners, preferences, tie_break)  # Find a single winner if multiple are there, by calling tie_breaking

def borda(preferences, tie_break):
    """borda function takes in a preference profile and an integer corresponding to an agent,
       and produces the winner according to the borda rule as its output."""

    length = len(preferences[1])

    score = {}
    for i in range(1, length + 1):
        score[i] = 0  # Assigning score 0 to each alternatives

    for value in preferences.values():
        for i, j in enumerate(value):
            score[j] += length - i - 1  # Assigning scores according to borda rule

    max_score = max(score.values())
    winners = find_winner(score, max_score)  # Find the winners with the maximum score by calling find_winner

    return tie_breaking(winners, preferences, tie_break)  # Find a single winner if multiple are there, by calling tie_breaking

def harmonic(preferences, tie_break):
    """harmonic function takes in a preference profile and an integer corresponding to an agent,
       and produces the winner according to the harmonic rule as its output."""

    length = len(preferences[1])
    score = {}
    for i in range(1, length + 1):
        score[i] = 0  # Assigning score 0 to each alternatives

    for value in preferences.values():
        for i, j in enumerate(value):
            score[length-j+1] += 1 / (length-i + 1)  # Assigning scores according to harmonic rule

    max_score = max(score.values())
    winners = find_winner(score, max_score)  # Find the winners with the maximum score by calling find_winner

    return tie_breaking(winners, preferences, tie_break)  # Find a single winner if multiple are there, by calling tie_breaking

def STV(preferences, tie_break):
    """STV function takes in a preference profile and an integer corresponding to an agent,
       and produces the winner according to the STV rule as its output."""

    preferences_copy = {}
    for key, value in preferences.items():
        preferences_copy[key] = value[:]  # created copy of original preferences

    while True:
        counts = {}
        for agent_preferences in preferences_copy.values():
            alternative = agent_preferences[0]
            if alternative in counts:
                counts[alternative] += 1  # Counts times each alternative came in the first position
            else:
                counts[alternative] = 1

        min_count = min(counts.values())
        least_freq = []
        for key, value in counts.items():
            if value == min_count:
                least_freq.append(key)  # Appending leat frequent  alernative in least_freq list

        final_preferences = {}
        for key, value in preferences_copy.items():
            final_preferences[key] = value[:]  # created copy of preferences_copy to get final preferences list

        for agent_preferences in preferences_copy.values():
            for x in least_freq:
                if x in agent_preferences:
                    agent_preferences.remove(x)  # Removing least frequent alternatives

        for key, value in preferences_copy.items():
            if len(value) > 0:
                preferences_copy[key]= value  # removing alternatives with zero values from consideration

        if len(counts.values()) == 1 or len(preferences_copy) <= 1:
            break  # breaking the loop if only one alternative left or all alternatives have the same count

    winners = list()
    for values in final_preferences.values():
        winners.append(values[0])  # getting possible winners from final_preferences list

    return tie_breaking(winners, preferences, tie_break)  # Find a single winner if multiple are there, by calling tie_breaking

def range_voting(values, tie_break):
    """range_voting function takes in a preference profile and an integer corresponding to an agent,
       and produces the winner according to the range_voting rule as its output."""

    value_list = list()
    for row in values.iter_rows(min_row=1, min_col=1, values_only=True):
        value_list.append(row)  # Appending each row values in value_list list

    preferences = {}
    for row_index, row in enumerate(value_list):
        agent_preferences = []
        for value in row:
            agent_preferences.append(float(value))  # appending each value to agent_preferences list

        preferences[row_index + 1] = agent_preferences  # Assgning each agent_preferences list to prefrences dictionary

    alternative_total = {}
    for i in range(1, (len(preferences[1])+1)):
        total_score = 0
        for value in preferences.values():
            total_score += value[i - 1]
        alternative_total[i] = total_score  # Calculating the total score for each alternative

    max_total = max(alternative_total.values())
    winners = find_winner(alternative_total, max_total)  # Find the winners with the maximum alternative_total by calling find_winner
    preference = generate_preferences(values)  # getting preference for tie_breaking function

    return tie_breaking(winners, preference, tie_break)  # Find a single winner if multiple are there, by calling tie_breaking

def tie_breaking(winners, preferences, tie_break):
    """tie_breaking is an extra function to find a singuler winner from multiple winners according to tie_break option."""

    if tie_break == "max":
        return max(winners)  # return a winner which is maximun number
    elif tie_break == "min":
        return min(winners)  # return a winner which is minimum number
    elif isinstance(tie_break, int) and tie_break in preferences:
        """if option is integer and also in preferences ordering then return the winner as
           one that agent(integer number) ranks the highest in his/her preference ordering."""

        for i in winners:
            max_index = -1
            max_value = None
            preferences_index = preferences[tie_break].index(i)

            if preferences_index > max_index:
                max_index = preferences_index
                max_value = i
        return max_value
    else:
        raise ValueError("Entered tie-breaking option is invalid!!")  # raise value error if integer is not in preferences

def find_winner(values, max):
    """find_winner is an extra function to find winners with maximum value from given dictionary"""

    winners = []
    for key, value in values.items():
        if value == max:
            winners.append(key)  # appending winners with maximum value
    return winners

def read_excel(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return sheet

while True:
      preferences = generate_preferences(read_excel("C:/Users/bonil/Downloads/voting.xlsx"))
      vector= [5,5,6,6]
      print(dictatorship(preferences, 2))
      print(scoring_rule(preferences, vector, "max"))
      print(plurality(preferences, "min"))
      print(veto(preferences, "max"))
      print(borda(preferences, "min"))
      print(harmonic(preferences, "max"))
      print(STV(preferences, "min"))
      print(range_voting(read_excel("C:/Users/bonil/Downloads/voting.xlsx"), "max"))
      break
